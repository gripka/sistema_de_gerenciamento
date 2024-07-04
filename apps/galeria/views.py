from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User, Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.messages import success
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from django.db.models import Q, Count
from django.db.models.functions import Lower, Substr, Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import View
from django.views.generic.edit import CreateView

import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime, timedelta
from .filters import GroupFilter
from .forms import (
    CadastroForm,
    EditarUsuarioForm,
    TransacaoForm,
    GroupForm,
    ModuloForm,
    FuncaoForm,
    ControleDeAcessoForm,
)
from .models import (
    Transacao, 
    Modulo, 
    UrlPermission, 
    User, 
    Group, 
    Permission
)


def index(request):
    if not request.user.is_authenticated:
        return redirect("usuarios:login")

    return render(request, "galeria/gerenciar_usuarios.html")


def meu_perfil(request):
    if not request.user.is_authenticated:
        return redirect("usuarios:login")
    return render(request, "galeria/meu_perfil.html")


@login_required
def gerenciar_usuarios(request):
    busca = request.GET.get("q", "")
    page_number = request.GET.get("page")

    usuarios = User.objects.annotate(
        primeira_letra=Lower(Substr(Coalesce("first_name", "username"), 1, 1))
    ).order_by("primeira_letra")

    if busca:
        usuarios = usuarios.filter(
            Q(username__icontains=busca)
            | Q(first_name__icontains=busca)
            | Q(last_name__icontains=busca)
            | Q(email__icontains=busca)
        )

    paginator = Paginator(usuarios, 7)
    page_obj = paginator.get_page(page_number)

    grupos = Group.objects.all()

    if request.method == "POST":
        usuario_id = request.POST.get("usuario_id")
        try:
            usuario = User.objects.get(id=usuario_id)
            usuario.delete()
            messages.success(request, "Usuário excluído com sucesso!")
        except User.DoesNotExist:
            messages.error(request, "Usuário não encontrado.")

        return redirect(
            "galeria:gerenciar_usuarios"
        )

    if request.method == "POST" and "cadastrar_usuario" in request.POST:
        form = CadastroForm(request.POST)
        if form.is_valid():
            print(form.cleaned_data) 
            usuario = form.save()
            print(usuario) 
            messages.success(request, "Novo usuário cadastrado com sucesso!")
            return redirect(
                f"{request.path}?page={page_obj.number}&q={busca}"
            )  
    else:
        form = CadastroForm()

    if request.method == "POST" and "editar_perfis" in request.POST:
        usuario_id = request.POST.get("usuario_id")
        perfis_selecionados = request.POST.getlist("perfis")

        usuario = get_object_or_404(User, pk=usuario_id)
        grupos_atuais = usuario.groups.all()

        for grupo in grupos_atuais:
            if str(grupo.id) not in perfis_selecionados:
                usuario.groups.remove(grupo)

        for perfil_id in perfis_selecionados:
            grupo = Group.objects.get(pk=perfil_id)
            usuario.groups.add(grupo)

        messages.success(request, "Perfis do usuário atualizados com sucesso!")
        return redirect(
            f"{request.path}?page={page_obj.number}&q={busca}"
        ) 

    return render(
        request,
        "galeria/gerenciar_usuarios.html",
        {
            "usuarios": page_obj,
            "grupos": grupos,
            "busca": busca,
            "form": form, 
        },
    )


class UserCreateView(CreateView):
    form_class = CadastroForm 
    template_name = "galeria/cadastro.html"

    def form_valid(self, form):
        user = form.save()
        success(
            self.request, "Novo usuário cadastrado com sucesso!"
        )  
        return render(
            self.request, self.template_name, {"form": self.form_class()}
        ) 

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))


@login_required
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    group_filter = GroupFilter(request.GET, queryset=Group.objects.all())

    if request.method == "POST":
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.groups.set(form.cleaned_data["groups"])
            usuario.save()
            messages.success(request, "Usuário editado com sucesso!")
        else:
            messages.error(request, "Erro ao editar usuário. Verifique os dados.")
    else:
        form = EditarUsuarioForm(instance=usuario)

    return render(
        request,
        "galeria/editar_usuario.html",
        {
            "form": form,
            "usuario": usuario,
            "filter": group_filter,
        },
    )


def transacoes(request):
    if not request.user.is_authenticated:
        return redirect("usuarios:login")

    transacoes = Transacao.objects.all()

    return render(request, "galeria/transacoes.html", {"transacoes": transacoes})


@login_required
def criar_transacao(request):
    if request.method == "POST":
        form = TransacaoForm(request.POST)
        if form.is_valid():
            transacao = form.save(commit=False)  
            transacao.save()
            form.save_m2m()  
            messages.success(request, "Transação criada com sucesso.")
            return redirect("galeria:editar_transacao", transacao_id=transacao.id)
        else:
            messages.error(request, "Nome já existente")
    else:
        form = TransacaoForm()

    permissoes = Permission.objects.all()
    permissoes_formatadas = []
    for permissao in permissoes:
        permissoes_formatadas.append(
            {
                "id": permissao.id,
                "name": permissao.name,
                "codename": permissao.codename,
                "checked": False,  
            }
        )

    return render(
        request,
        "galeria/criar_transacao.html",
        {
            "form": form,
            "permissoes": permissoes_formatadas,
        },
    )


@login_required
def editar_transacao(request, transacao_id):
    transacao = get_object_or_404(Transacao, id=transacao_id)
    permissoes = Permission.objects.all()

    if request.method == "POST":
        form = TransacaoForm(request.POST, instance=transacao)
        if form.is_valid():
            transacao = form.save(commit=False)
            transacao.permissoes.set(form.cleaned_data["permissoes"])
            transacao.save()
            messages.success(request, "Transação atualizada com sucesso.")
            return redirect("galeria:editar_transacao", transacao_id=transacao.id)
        else:
            messages.error(request, "Nome já existente")

    else:
        form = TransacaoForm(instance=transacao)

    permissoes_formatadas = []
    for permissao in permissoes:
        checked = permissao in transacao.permissoes.all()
        permissoes_formatadas.append(
            {
                "id": permissao.id,
                "name": permissao.name,
                "codename": permissao.codename,
                "checked": checked,
            }
        )

    return render(
        request,
        "galeria/editar_transacao.html",
        {
            "form": form,
            "transacao": transacao,
            "permissoes": permissoes_formatadas,
        },
    )


def excluir_transacao(request, pk):
    transacao = Transacao.objects.get(pk=pk)
    if request.method == "POST":
        transacao.delete()
        messages.success(request, "Transação excluída com sucesso!")
    return redirect("galeria:transacoes")


@login_required
def gestao_de_perfis(request):
    grupos = Group.objects.all() 
    return render(
        request,
        "galeria/gestao_de_perfis.html",
        {"grupos": grupos, "teste": "Olá, mundo!"},
    )


@login_required
def excluir_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)

    if request.method == "POST":
        usuario.delete()
        return redirect("galeria:gerenciar_usuarios")


def buscar_grupos(request):
    query = request.GET.get("q", "")
    grupos = Group.objects.filter(Q(name__icontains=query))
    data = [{"id": grupo.id, "name": grupo.name} for grupo in grupos]
    return JsonResponse(data, safe=False)


@login_required
def criar_grupo(request):
    modulos = Modulo.objects.all()
    transacoes_disponiveis = Transacao.objects.all()

    if request.method == "POST":
        form = GroupForm(request.POST)
        if form.is_valid():
            grupo = form.save()

            transacoes_ids = request.POST.getlist("transacoes", [])
            if transacoes_ids:
                transacoes = Transacao.objects.filter(id__in=transacoes_ids)
                grupo.transacoes.set(transacoes)

            modulos_ids = request.POST.getlist("modulos", [])
            if modulos_ids:
                modulos = Modulo.objects.filter(id__in=modulos_ids)
                grupo.modulos.set(modulos)  

            messages.success(request, "Grupo criado com sucesso!")
            return redirect("galeria:gestao_de_perfis")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)

    else:
        form = GroupForm()
        
    context = {
        "form": form,
        "modulos": modulos,
        "transacoes_disponiveis": transacoes_disponiveis,
    }
    return render(request, "galeria/criar_grupo.html", context)


@login_required
def editar_grupo(request, grupo_id):
    grupo = get_object_or_404(Group, pk=grupo_id)
    modulos_disponiveis = Modulo.objects.all()
    transacoes_disponiveis = Transacao.objects.all()

    if request.method == 'POST':
        form = GroupForm(request.POST, instance=grupo)
        if form.is_valid():
            grupo = form.save(commit=False)

            transacoes_ids = request.POST.getlist("transacoes", [])
            transacoes = Transacao.objects.filter(id__in=transacoes_ids)
            grupo.transacoes.set(transacoes)

            modulos_ids = request.POST.getlist("modulos", [])
            modulos = Modulo.objects.filter(id__in=modulos_ids)
            grupo.modulos.set(modulos)

            permissoes_a_remover = []
            for permissao in grupo.permissions.all():
                if not permissao.modulos.filter(id__in=modulos_ids).exists() and not permissao.transacoes.filter(id__in=transacoes_ids).exists():
                    permissoes_a_remover.append(permissao)

            grupo.permissions.remove(*permissoes_a_remover)

            novas_permissoes = set()
            for modulo in modulos:
                novas_permissoes.update(modulo.permissions.all())
            for transacao in transacoes:
                novas_permissoes.update(transacao.permissoes.all())

            grupo.permissions.add(*novas_permissoes)

            grupo.save()

            messages.success(request, 'Perfil editado com sucesso!')
            return redirect('galeria:gestao_de_perfis')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = GroupForm(instance=grupo)

    modulos_do_grupo = grupo.modulos.all()
    transacoes_do_grupo = grupo.transacoes.all()

    context = {
        'form': form,
        'modulos_disponiveis': modulos_disponiveis,
        'transacoes_disponiveis': transacoes_disponiveis,
        'modulos_do_grupo': modulos_do_grupo,
        'transacoes_do_grupo': transacoes_do_grupo,
    }
    return render(request, 'galeria/editar_grupo.html', context)  


@login_required
def excluir_grupo(request, pk):
    grupo = get_object_or_404(Group, pk=pk)

    if request.method == "POST":
        grupo.delete()
        messages.success(request, "Perfil excluído com sucesso!")
        return redirect("galeria:gestao_de_perfis")
    else:
        return render(
            request, "galeria/confirmar_exclusao_grupo.html", {"grupo": grupo}
        )


@login_required
def listar_modulos(request):
    query = request.GET.get("q")
    if query:
        modulos = Modulo.objects.filter(
            Q(nome__icontains=query) | Q(descricao__icontains=query)
        )
    else:
        modulos = Modulo.objects.all()
    return render(request, "galeria/modulos.html", {"modulos": modulos})


@login_required
def criar_modulo(request):
    if request.method == "POST":
        form = ModuloForm(request.POST)
        if form.is_valid():
            modulo = form.save()
            transacoes_ids = request.POST.getlist("transacoes")
            permissoes_ids = request.POST.getlist("permissions")
            
            for transacao_id in transacoes_ids:
                transacao = Transacao.objects.get(id=transacao_id)
                transacao.modulo = modulo
                transacao.permissoes.set(permissoes_ids)  
                transacao.save()
            modulo.propagate_permissions() 
            messages.success(request, "Módulo criado com sucesso!")
            return redirect("galeria:listar_modulos")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = ModuloForm()

    transacoes = Transacao.objects.all()
    return render(
        request, "galeria/criar_modulo.html", {"form": form, "transacoes": transacoes}
    )


@login_required
def editar_modulo(request, pk):
    modulo = get_object_or_404(Modulo, pk=pk)
    error_message = None

    if request.method == "POST":
        form = ModuloForm(request.POST, instance=modulo)
        if form.is_valid():
            modulo = form.save()
            transacoes_ids = request.POST.getlist("transacoes")
            permissions_ids = request.POST.getlist("permissions")

            for transacao_id in transacoes_ids:
                transacao = Transacao.objects.get(id=transacao_id)
                transacao.modulo = modulo
                transacao.permissoes.set(permissions_ids)  
                transacao.save()

            modulo.permissions.set(permissions_ids)
            messages.success(request, "Módulo editado com sucesso!")
            return redirect("galeria:listar_modulos")
        else:
            if "nome" in form.errors:
                messages.error(request, form.errors["nome"][0])
    else:
        form = ModuloForm(instance=modulo)

    transacoes = Transacao.objects.all()
    for transacao in transacoes:
        transacao.checked = transacao in modulo.transacoes.all()

    return render(
        request,
        "galeria/editar_modulo.html",
        {"form": form, "modulo": modulo, "transacoes": transacoes},
    )


@login_required
def excluir_modulo(request, pk):
    modulo = get_object_or_404(Modulo, pk=pk)
    if request.method == "POST":
        modulo.delete()
        messages.success(request, "Módulo excluído com sucesso!")
        return redirect("galeria:listar_modulos")
    return render(request, "galeria/confirmar_exclusao_modulo.html", {"modulo": modulo})


@login_required
def listar_funcoes(request):
    funcoes = Permission.objects.all()
    return render(request, "galeria/funcoes.html", {"funcoes": funcoes})


@login_required
def criar_funcao(request):
    user = request.user
    print(user.get_all_permissions())

    if request.method == "POST":
        form = FuncaoForm(request.POST)
        if form.is_valid():
            funcao = form.save(commit=False)

            content_type = ContentType.objects.get(app_label='auth', model='permission')
            funcao.content_type = content_type
            funcao.save()

            messages.success(request, "Função criada com sucesso!")
            return redirect("galeria:listar_funcoes")
    else:
        form = FuncaoForm()

    return render(request, "galeria/criar_funcao.html", {"form": form})


@login_required
def editar_funcao(request, pk):
    funcao = get_object_or_404(Permission, pk=pk)

    if request.method == "POST":
        form = FuncaoForm(request.POST, instance=funcao)
        if form.is_valid():
            funcao = form.save(commit=False)

            content_type = ContentType.objects.get(app_label='auth', model='permission')
            funcao.content_type = content_type
            funcao.save()

            messages.success(request, "Função editada com sucesso!")
            return redirect("galeria:listar_funcoes")
    else:
        form = FuncaoForm(instance=funcao)

    return render(request, "galeria/editar_funcao.html", {"form": form, "funcao": funcao})


@login_required
def excluir_funcao(request, pk):
    funcao = get_object_or_404(Permission, pk=pk)
    if request.method == "POST":
        funcao.delete()
        messages.success(request, "Função excluída com sucesso!")
        return redirect("galeria:listar_funcoes")


@login_required
def relatorios(request):
#    if not request.user.has_perm('auth.VSRS'):
#        raise PermissionDenied("Você não tem permissão para acessar esta página.")
    usuarios = User.objects.all()
    perfis = Group.objects.all()
    modulos = Modulo.objects.all()
    transacoes = Transacao.objects.all()
    permissoes = Permission.objects.all()

    context = {
        "usuarios": usuarios,
        "perfis": perfis,
        "modulos": modulos,
        "transacoes": transacoes,
        "permissoes": permissoes,
    }
    return render(request, "galeria/relatorios.html", context)


@login_required
@permission_required('auth.VSRS', raise_exception=True)
def exportar_relatorios(request):
    usuarios = User.objects.all()
    perfis = Group.objects.all()
    modulos = Modulo.objects.all()
    transacoes = Transacao.objects.all()
    permissoes = Permission.objects.all()

    hoje = datetime.now().date()
    data_30_dias_atras = hoje - timedelta(days=30)
    data_7_dias_atras = hoje - timedelta(days=7)

    usuarios_30_dias = usuarios.filter(date_joined__date__gte=data_30_dias_atras)
    usuarios_7_dias = usuarios.filter(date_joined__date__gte=data_7_dias_atras)

    wb = Workbook()

    def add_to_sheet(sheet, headers, data, is_usuarios_sheet=False):
        font_bold = Font(bold=True)
        fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        border_thin = Border(left=Side(style="thin"), 
                            right=Side(style="thin"), 
                            top=Side(style="thin"), 
                            bottom=Side(style="thin"))

        if is_usuarios_sheet:
            sheet.append(["", "", "", "", "", "", "", ""])
            sheet.append(["Total de Usuários", "", len(usuarios), "", "", ""])
            sheet.append(["Criados nos últimos 30 dias", "", len(usuarios_30_dias), "", "", ""])
            sheet.append(["Criados nos últimos 7 dias", "", len(usuarios_7_dias), "", "", ""])
            sheet.append(["", "", "", "", "", "", "", ""]) 

            for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.min_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    if cell.value in headers: 
                        cell.font = font_bold
                    cell.fill = fill_grey
                    cell.border = border_thin

        sheet.append(headers)

        if is_usuarios_sheet:
            for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.font = font_bold

        for row in data:
            sheet.append(row)

        for row in sheet.iter_rows(min_row=sheet.min_row + 1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border_thin

    def formatar_usuarios():
        usuarios_sheet = wb.active
        usuarios_sheet.title = "Usuários Cadastrados"
        usuarios_headers = [
            "Usuário", "Nome", "Sobrenome", "Email", "Data de Cadastro", "Último Login", "Ativo", "Grupos"
        ]
        usuarios_data = [
            [
                usuario.username,
                usuario.first_name,
                usuario.last_name,
                usuario.email,
                usuario.date_joined.strftime("%Y-%m-%d %H:%M:%S"),
                usuario.last_login.strftime("%Y-%m-%d %H:%M:%S") if usuario.last_login else "",
                "Sim" if usuario.is_active else "Não",
                ", ".join([grupo.name for grupo in usuario.groups.all()]),
            ]
            for usuario in usuarios
        ]

        add_to_sheet(usuarios_sheet, usuarios_headers, usuarios_data, is_usuarios_sheet=True)  

        for cell in usuarios_sheet[6]:  
            cell.font = Font(bold=True)

    def formatar_perfis():
        perfis_sheet = wb.create_sheet(title="Perfis de Usuários")
        perfis_headers = ["Nome do Perfil", "Funções", "Módulos", "Transações"]
        perfis_data = [
            [
                perfil.name,
                ", ".join([permissao.name for permissao in perfil.permissions.all()]),
                ", ".join([modulo.nome for modulo in perfil.modulos.all()]),
                ", ".join([transacao.nome for transacao in perfil.transacoes.all()]),
            ]
            for perfil in perfis
        ]
        add_to_sheet(perfis_sheet, perfis_headers, perfis_data)

    def formatar_modulos():
        modulos_sheet = wb.create_sheet(title="Lista de Módulos")
        modulos_headers = ["Nome do Módulo", "Descrição", "Transações Associadas"]
        modulos_data = [
            [
                modulo.nome,
                modulo.descricao,
                "\n".join([f"{transacao.nome} - {transacao.descricao}" for transacao in modulo.transacoes.all()])
            ]
            for modulo in modulos
        ]
        add_to_sheet(modulos_sheet, modulos_headers, modulos_data)

    def formatar_transacoes():
        transacoes_sheet = wb.create_sheet(title="Lista de Transações")
        transacoes_headers = ["Nome da Transação", "Descrição", "Funções"]
        transacoes_data = [
            [
                transacao.nome,
                transacao.descricao,
                ", ".join([permissao.name for permissao in transacao.permissoes.all()]),
            ]
            for transacao in transacoes
        ]
        add_to_sheet(transacoes_sheet, transacoes_headers, transacoes_data)

    def formatar_permissoes():
        permissoes_sheet = wb.create_sheet(title="Funções Cadastradas")
        permissoes_headers = ["Nome", "Código"]
        permissoes_data = [
            [permissao.name, permissao.codename]
            for permissao in permissoes
        ]
        add_to_sheet(permissoes_sheet, permissoes_headers, permissoes_data)

    formatar_usuarios()
    formatar_perfis()
    formatar_modulos()
    formatar_transacoes()
    formatar_permissoes()

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorios.xlsx"'
    wb.save(response)

    return response


@login_required
def usuarios_cadastrados(request):
    search_query = request.GET.get('search', '')
    group_filter = request.GET.get('group', '')

    usuarios = User.objects.all()
    if search_query:
        usuarios = usuarios.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    if group_filter:
        usuarios = usuarios.filter(groups__id=group_filter)

    total_usuarios = usuarios.count()
    usuarios_ultimos_30_dias = usuarios.filter(date_joined__gte=datetime.now()-timedelta(days=30)).count()

    grupos = Group.objects.all()

    return render(request, "galeria/usuarios_cadastrados.html", {
        "usuarios": usuarios,
        "total_usuarios": total_usuarios,
        "usuarios_ultimos_30_dias": usuarios_ultimos_30_dias,
        "grupos": grupos
    })


@login_required
def perfis_usuarios(request):
    perfis = Group.objects.all()
    return render(request, "galeria/perfis_usuarios.html", {"perfis": perfis})


@login_required
def lista_modulos(request):
    modulos = Modulo.objects.all()
    return render(request, "galeria/lista_modulos.html", {"modulos": modulos})


@login_required
def lista_transacoes(request):
    transacoes = Transacao.objects.all()
    return render(request, "galeria/lista_transacoes.html", {"transacoes": transacoes})


@login_required
def funcoes_cadastradas(request):
    funcoes = Permission.objects.all()
    return render(request, "galeria/funcoes_cadastradas.html", {"funcoes": funcoes})


@login_required
@permission_required('auth.CTRA')
def controle_de_acesso(request):
    if request.method == 'POST':
        form = ControleDeAcessoForm(request.POST)
        if form.is_valid():
            form.save()  
            messages.success(request, 'Permissões atualizadas com sucesso!')
            return redirect('galeria:controle_de_acesso')  

    url_permissions = UrlPermission.objects.all()

    url_to_permission = {url_permission.url: url_permission.permissions.first().id if url_permission.permissions.exists() else None for url_permission in url_permissions}

    form = ControleDeAcessoForm(initial=url_to_permission)

    context = {'form': form}
    return render(request, 'galeria/controle_de_acesso.html', context)


def acesso_negado(request):
    return render(request, 'galeria/acesso_negado.html')